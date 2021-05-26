from openpyxl import load_workbook
import getpass

## 불러올 엑셀 파일 설정

# 현재 컴퓨터의 유저명 들고오기
userName = getpass.getuser()

# oneDrive 주소 가져오기
path1 = "C:/Users/"
path2 = "/OneDrive/AI_List.xlsx"

path = path1 + str(userName) + path2

# data_only=True : 수식없이 값만 가져오게 설정
# 내가 저장한 AI_List 들고오기
load_wb = load_workbook(path, data_only=True)

# 불러올 시트 설정
load_sheet = load_wb['평균선-해외']

StockList = []  # 인공지능 종목 리스트
webPageList = []  # 인공지능 추천가

## *** 엑셀 데이터 저장
def saveExcell(load_wb):
    # 엑셀이 열려있을 때 저장하려고 할 경우 에러 메시지 출력
    try:
        excell_file = load_wb
        excell_file.save(path)
    except PermissionError:
        print("열려있는 AI_List.xlsx 엑셀을 닫으세요")

# 같은 이름의 종목명 있는지 찾아서 삭제하기 - averageStock에선 사용하지 않음
def delete_stock(stock):
    for i in range(2, 200):
        # 'B2'부터 'B100'까지 읽기
        stock_name = load_sheet.cell(i, 2).value
        # 셀 비었으면 그대로 종료
        if stock_name == None:
            break
        # 종목명이 일치하면 행 삭제
        if stock_name == stock:
            load_sheet.delete_rows(i)
            # 행 삭제 후 저장하기
            saveExcell(load_wb)

def read_stock():
    stock_list = []
    for i in range(11, 100):
        # 'B10'부터 아래로 읽기
        stock_name = load_sheet.cell(i, 2).value
        if stock_name == None:
            break
        stock_list.append(stock_name)

    return stock_list

def read_code():
    code_list = []
    for i in range(11, 100):
        # 'A10'부터 아래로 읽기
        stock_name = load_sheet.cell(i, 1).value
        if stock_name is None:
            break
        code_list.append(stock_name)

    return code_list

def write_excel(row, average5, average20, average60, average120, text, result):
    # 각각의 표에 데이터 넣기
    load_sheet.cell(row, 3).value = average5
    load_sheet.cell(row, 4).value = average20
    load_sheet.cell(row, 5).value = average60
    load_sheet.cell(row, 6).value = average120
    load_sheet.cell(row, 7).value = str(result) + "%"
    load_sheet.cell(row, 8).value = text

    saveExcell(load_wb)