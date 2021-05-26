import yfinance as yf
import average_Stock_excel
import datetime
import average_Stock_excel
from openpyxl import load_workbook
import getpass
import datetime

def findAndWrite():
    stock_list = average_Stock_excel.read_stock()
    code_list = average_Stock_excel.read_code()

    # 현재 컴퓨터의 유저명 들고오기
    userName = getpass.getuser()

    # oneDrive 주소 가져오기
    path1 = "C:/Users/"
    path2 = "/OneDrive/AI_List.xlsx"

    path = path1 + str(userName) + path2
    load_wb = load_workbook(path, data_only=True)
    load_sheet = load_wb['평균선']

    # 날짜 가져오기
    today = datetime.datetime.now()
    start_day = today - datetime.timedelta(days=+1)

    # 날짜 양식 변환(이 부분은 krx 파이썬이랑 다름)
    today = today.strftime("%Y-%m-%d")
    start_day = start_day.strftime("%Y-%m-%d")
    # 프로그램 실시 날짜 기록
    load_sheet.cell(3, 2).value = str(today)

    # 엑셀 기록을 11번 row 부터 시작함
    i = 11

    for code in code_list:
        print(code)
        data = yf.Ticker(code)
        # 해당 종목의 해당 기간동안의 종가 데이터만 가져오기
        # Valid periods: 1d,5d,1mo,3mo,6mo,1y,2y,5y,10y,ytd,max
        hist = data.history(period="1y")[['Close']]

        # 제일 최근 날짜의 종가 가져오기
        hist1 = hist.tail(1)
        recent_price_2 = hist1.values.tolist()
        recent_price_3 = sum(recent_price_2, [])
        recent_price = recent_price_3[0]

        # 5일 평균 구하기
        # 가져온 데이터는 최근 데이터일수록 아래에 가며 스택 형식으로 출력된다
        # -> 그래서 아래의 5개만 들고온다
        hist5 = hist.tail(5)
        # 가져온 종목은 데이터 프레임이기 때문에 연산을 위해 리스트로 바꿔주기
        hist_list = hist5.values.tolist()
        # 이상하게 2차원으로 들어가 있으므로 1차원 리스트로 바꾸기
        hist_list = sum(hist_list, [])
        average5 = round(sum(hist_list) / 5, 2)

        # 20일 평균 구하기
        hist20 = hist.tail(20)
        hist_list = hist20.values.tolist()
        hist_list = sum(hist_list, [])
        average20 = round(sum(hist_list) / 20, 2)

        # 60일 평균 구하기
        hist60 = hist.tail(60)
        hist_list = hist60.values.tolist()
        hist_list = sum(hist_list, [])
        average60 = round(sum(hist_list) / 60, 2)

        # 120일 평균 구하기
        hist120 = hist.tail(120)
        hist_list = hist120.values.tolist()
        hist_list = sum(hist_list, [])
        average120 = round(sum(hist_list) / 120, 2)

        # print(average5)
        # print(average20)
        # print(average60)
        # print(average120)

        # 제일 최근 종가와 각각의 평균선 값 비교하기
        text = ""
        result = 0
        # 120일 선보다 낮음 & 5일 위
        if average120 > recent_price > average5:
            text = "120일 보다 낮음 & 5일 위"
            result = round((recent_price - average120) / average120 * 100, 1)
        # 120일 보다 낮을 때
        elif recent_price < average120:
            text = "120일 보다 낮음"
            result = round((recent_price - average120) / average120 * 100, 1)
        # 60일 보다 낮음 & 5일 위
        elif average60 > recent_price > average5:
            text = "60일 보다 낮음 & 5일 위"
            result = round((recent_price - average60) / average60 * 100, 1)
        # 60일 보다 낮음
        elif recent_price < average60:
            text = "60일 보다 낮음"
            result = round((recent_price - average60) / average60 * 100, 1)
        # 20일 보다 낮음 & 5일 위
        elif average20 > recent_price > average20:
            text = "20일 보다 낮음 & 5일 위"
            result = round((recent_price - average20) / average20 * 100, 1)
        # 20일 보다 낮음
        elif recent_price < average20:
            text = "20일 보다 낮음"
            result = round((recent_price - average20) / average20 * 100, 1)
        # 5일 보다 낮음
        elif recent_price < average5:
            text = "5일 보다 낮음"
            result = round((recent_price - average5) / average5 * 100, 1)
        else:
            text = ""
            result = ""

        average_Stock_excel.write_excel(i, average5, average20, average60, average120, text, result)

        i += 1






